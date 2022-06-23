from glob import glob
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import numpy as np
import os

# SET THE INSTANCE NUMBER
instance_number = 24

# INSERT HERE YOUR SOLUTION
solution = [13, 0, 13, 3, 0, 0, 0, 5, 16, 0, 0, 11, 6, 22, 12, 7, 9, 7, 0, 19, 9, 18, 12, 17, 6, 0, 3, 0, 15, 7, 9, 0, 15, 10, 15, 18, 19, 0]


# re-organized solution in e.g. [5, 5, 5, 0, 0, 5, 0, 0] -> [[5, 5], [5, 0], [0, 5], [0, 0]]
solution = [solution[i: i + 2] for i in range(0, len(solution), 2)]

COLORS = [
    '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF',  # 0-4
    '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF',  # 5-9
    '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF',  # 10-14
    '0000FFFF', '00800000', '00008000', '00000080', '00808000',  # 15-19
    '00800080', '00008080', '00C0C0C0', '00808080', '009999FF',  # 20-24
    '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080',  # 25-29
    '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00',  # 30-34
    '0000FFFF', '00800080', '00800000', '00008080', '000000FF',  # 35-39
    '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF',  # 40-44
    '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC',  # 45-49
    '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699',  # 50-54
    '00969696', '00003366', '00339966', '00003300', '00333300',  # 55-59
    '00993300', '00993366', '00333399', '00333333',  # 60-63
]

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

instance_path = './instances_txt/ins-' + str(instance_number) + '.txt'
# open the instance file
with open(instance_path, 'r') as f:
    width = int(f.readline())
    n = int(f.readline())

    # reading the shapes
    shapes = []
    for shape in f:
        x, y = shape.split(" ")
        x = int(x)
        y = int(y)
        shapes.append((x, y))

    # creating excel sheet
    workbook = Workbook()
    sheet = workbook.active

    worst_max_heigth = np.sum([y[1] for y in shapes])

    # setting the width of width column and creating the grid with borders
    for i in range(1,width+1):
        sheet.column_dimensions[get_column_letter(i)].width = 3
        for j in range(1, worst_max_heigth + 1):
            sheet.cell(row = j, column= i).border = thin_border
            sheet.cell(row = j, column= i).alignment = Alignment(horizontal='center')

    # for each solution draw a rectangle
    for i in range(0, len(solution)):
        color = PatternFill(start_color=COLORS[i], end_color=COLORS[i], fill_type='solid')

        block_shape = shapes[i]
        block_coord = solution[i]

        starting_x = block_coord[0] + 1
        starting_y = block_coord[1] + 1

        ending_x = starting_x + int(block_shape[0])
        ending_y = starting_y + int(block_shape[1])

        for x in range(starting_x, ending_x):
            for y in range(starting_y, ending_y):
                sheet.cell(row=y, column=x).fill = color
                sheet.cell(row=y, column=x).border = thin_border
                sheet.cell(row=y, column=x).value = i

    if not os.path.exists('./displayed_solutions'):
        os.mkdir('./displayed_solutions')

    workbook.save(filename='./displayed_solutions/instance_' + str(instance_number) + ".xlsx")



